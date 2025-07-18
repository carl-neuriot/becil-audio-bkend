import time
import os
import numpy as np
import pandas as pd
from datetime import datetime
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.ext.declarative import declarative_base
from models import Ad, Song, Broadcast, AdDetectionResult, ExcelReports
from io import BytesIO
from fastapi.responses import StreamingResponse
import soundfile as sf
from pydub import AudioSegment
from fastapi import HTTPException
import logging
import traceback
import io
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from s3 import download_file_from_s3
import schedule
from audioutils import (
    seconds_to_standard_time,
    extract_brand_name,
    load_audio,
    find_matches_improved
)
from crud import get_broadcast

Base = declarative_base()


class EnhancedRadioRecordingManager:
    def __init__(self, db_path="radio_ad_detection.db"):
        self.db_path = db_path
        self.engine = create_engine(f"sqlite:///{db_path}")
        Base.metadata.create_all(self.engine)
        Session = sessionmaker(bind=self.engine)
        self.Session = Session

    def get_ad_id_by_filename(self, filename):
        """Get ad ID from the ads table by filename"""
        session = self.Session()
        try:
            ad = session.query(Ad).filter(Ad.advertisement == filename).first()
            return ad.id if ad else -1
        except Exception as e:
            print(f"Error getting ad ID for {filename}: {e}")
            return -1
        finally:
            session.close()

    def get_broadcast_by_id(self, broadcast_id):
        session = self.Session()
        try:
            broadcast = (
                session.query(Broadcast).filter(
                    Broadcast.id == broadcast_id).first()
            )
            return broadcast
        except Exception as e:
            print(f"Error getting broadcast for id {broadcast_id}: {e}")
            return -1
        finally:
            session.close()

    def get_broadcast_id_by_filename(self, filename):
        """Get broadcast ID from the broadcasts table by filename"""
        session = self.Session()
        try:
            broadcast = (
                session.query(Broadcast)
                .filter(Broadcast.broadcast_recording == filename)
                .first()
            )
            return broadcast.id if broadcast else -1
        except Exception as e:
            print(f"Error getting broadcast ID for {filename}: {e}")
            return -1
        finally:
            session.close()

    def _find_gaps_in_broadcast(self, broadcast_id: int, session):
        """
        Return list of (gap_start, gap_end) seconds where no ads/songs/speech exist.
        """
        results = (
            session.query(AdDetectionResult)
            .filter(AdDetectionResult.broadcast_id == broadcast_id)
            .order_by(AdDetectionResult.start_time_seconds)
            .all()
        )

        gaps = []
        last_end = 0.0

        for result in results:
            if result.start_time_seconds - last_end >= 1.0:
                gaps.append((last_end, result.start_time_seconds))
            last_end = max(last_end, result.end_time_seconds)

        broadcast = session.query(Broadcast).get(broadcast_id)
        broadcast_duration = (
            broadcast.duration if broadcast and broadcast.duration else last_end + 60
        )

        if broadcast_duration - last_end >= 1.0:
            gaps.append((last_end, broadcast_duration))

        return gaps

    def populate_ads_from_folder(self, ad_masters_folder="ad_masters"):
        """Populate ads table from ad_masters folder"""
        session = self.Session()
        try:
            print(f" Scanning {ad_masters_folder} for advertisement files...")

            if not os.path.exists(ad_masters_folder):
                print(f"✗ Folder {ad_masters_folder} does not exist")
                return 0

            added_count = 0
            updated_count = 0

            for filename in os.listdir(ad_masters_folder):
                if filename.endswith((".wav", ".mp3")):
                    filepath = os.path.join(ad_masters_folder, filename)

                    existing_ad = (
                        session.query(Ad).filter(
                            Ad.advertisement == filename).first()
                    )

                    duration_seconds = None
                    audio, sr = load_audio(filepath)
                    if audio is not None:
                        duration_seconds = int(len(audio) / sr)

                    brand_name = extract_brand_name(filename)

                    if existing_ad:
                        existing_ad.brand = brand_name
                        existing_ad.duration = duration_seconds
                        existing_ad.status = "Active"
                        updated_count += 1
                    else:

                        new_ad = Ad(
                            brand=brand_name,
                            advertisement=filename,
                            duration=duration_seconds,
                            status="Active",
                            upload_date=datetime.now(),
                        )
                        session.add(new_ad)
                        added_count += 1

            session.commit()
            print(" Ad table updated:")
            print(f"  - New ads added: {added_count}")
            print(f"  - Existing ads updated: {updated_count}")

            return added_count + updated_count

        except Exception as e:
            session.rollback()
            print(f"✗ Error populating ads table: {e}")
            return 0
        finally:
            session.close()

    def add_broadcast(
        self,
        broadcast_recording,
        radio_station=None,
        duration=None,
        broadcast_date=None,
        status="Pending",
    ):
        """Add a new broadcast to the database"""
        session = self.Session()
        try:

            existing = (
                session.query(Broadcast)
                .filter(Broadcast.broadcast_recording == broadcast_recording)
                .first()
            )

            if existing:
                print(
                    f" Broadcast {broadcast_recording} already exists in database (ID: {existing.id})"
                )
                return existing.id

            if broadcast_date is None:
                broadcast_date = datetime.now()

            new_broadcast = Broadcast(
                radio_station=radio_station,
                broadcast_recording=broadcast_recording,
                duration=duration,
                broadcast_date=broadcast_date,
                status=status,
            )

            session.add(new_broadcast)
            session.commit()

            print(
                f" Added broadcast: {broadcast_recording} (ID: {new_broadcast.id})")
            return new_broadcast.id

        except Exception as e:
            session.rollback()
            print(f" Error adding broadcast: {e}")
            return None
        finally:
            session.close()

    def update_broadcast_status(self, broadcast_id, status):
        """Update broadcast processing status"""
        session = self.Session()
        try:
            broadcast = (
                session.query(Broadcast).filter(
                    Broadcast.id == broadcast_id).first()
            )

            if broadcast:
                broadcast.status = status
                session.commit()
                print(f" Updated {broadcast_id} status to: {status}")
                return True
            else:
                print(f" Broadcast {broadcast_id} not found")
                return False

        except Exception as e:
            session.rollback()
            print(f" Error updating broadcast status: {e}")
            return False
        finally:
            session.close()

    def save_detection_results(self, broadcast_id, matches_dict, radio_filename):
        """Save detection results to database for a specific radio recording"""
        session = self.Session()

        try:

            if broadcast_id == -1:
                print(
                    f"⚠ Warning: Broadcast {radio_filename} not found in database")
                return 0

            session.query(AdDetectionResult).filter(
                AdDetectionResult.broadcast_id == broadcast_id
            ).delete()

            all_matches = []
            for master_name, matches in matches_dict.items():

                ad_id = self.get_ad_id_by_filename(master_name)

                for match in matches:
                    all_matches.append(
                        {
                            "master_name": master_name,
                            "ad_id": ad_id,
                            "start_time": match["start_time"],
                            "end_time": match["end_time"],
                            "duration": match["duration"],
                            "correlation": match["correlation"],
                            "overlap_duration": match.get(
                                "overlap_duration", match["duration"]
                            ),
                            "raw_correlation": match.get(
                                "raw_correlation", match["correlation"]
                            ),
                            "mfcc_correlation": match.get(
                                "mfcc_correlation", match["correlation"]
                            ),
                            "clip_type": "ad",
                        }
                    )

            all_matches.sort(key=lambda x: x["start_time"])
            final_matches = self._filter_overlapping_matches(all_matches)

            for match in final_matches:
                brand_name = extract_brand_name(match["master_name"])
                description = os.path.splitext(match["master_name"])[0]

                db_record = AdDetectionResult(
                    brand=brand_name,
                    description=description,
                    start_time_seconds=match["start_time"],
                    end_time_seconds=match["end_time"],
                    duration_seconds=match["duration"],
                    correlation_score=match["correlation"],
                    raw_correlation=match["raw_correlation"],
                    mfcc_correlation=match["mfcc_correlation"],
                    overlap_duration=match["overlap_duration"],
                    total_matches_found=len(final_matches),
                    ad_id=match["ad_id"],
                    broadcast_id=broadcast_id,
                    clip_type=match["clip_type"],
                    detection_timestamp=datetime.now(),
                )
                session.add(db_record)

            session.commit()
            print(
                f" Saved {len(final_matches)} detection results for broadcast id {broadcast_id}")

            self.update_broadcast_status(broadcast_id, "Completed")

            self._generate_and_store_excel(broadcast_id, final_matches)

            return len(final_matches)

        except Exception as e:
            session.rollback()
            print(f" Error saving results: {e}")
            return 0
        finally:
            session.close()

    def _filter_overlapping_matches(self, all_matches):
        """Apply filtering logic for overlapping matches"""
        final_matches = []
        for match in all_matches:
            should_keep = True

            for i, existing_match in enumerate(final_matches):
                overlap_start = max(
                    match["start_time"], existing_match["start_time"])
                overlap_end = min(match["end_time"],
                                  existing_match["end_time"])
                overlap_duration = max(0, overlap_end - overlap_start)

                min_duration = min(match["duration"],
                                   existing_match["duration"])

                if overlap_duration > (0.4 * min_duration):
                    current_score = match["correlation"]
                    existing_score = existing_match["correlation"]

                    if current_score > existing_score:
                        final_matches[i] = match
                        should_keep = False
                        break
                    else:
                        should_keep = False
                        break

            if should_keep:
                final_matches.append(match)

        return final_matches

    def _generate_and_store_excel(self, broadcast_id, final_matches):
        """Generate Excel and store in database, including empty slots."""
        session = self.Session()
        try:
            broadcast = (
                session.query(Broadcast).filter(
                    Broadcast.id == broadcast_id).first()
            )
            if not broadcast:
                print(f" Broadcast with ID {broadcast_id} not found")
                return

            output = BytesIO()

            all_rows = []
            last_end = 0.0

            # Sort matches by start time to correctly identify gaps
            final_matches.sort(key=lambda x: x['start_time'])

            for match in final_matches:
                start_time = match['start_time']

                # Add a gap if there is one of at least 1 second
                if start_time - last_end >= 1.0:
                    gap_start = last_end
                    gap_end = start_time
                    duration = gap_end - gap_start
                    all_rows.append({
                        "Brand": "Empty",
                        "Description": "Empty",
                        "Start Time (HH:MM:SS)": seconds_to_standard_time(round(gap_start)),
                        "End Time (HH:MM:SS)": seconds_to_standard_time(round(gap_end)),
                        "Ad Duration (HH:MM:SS)": seconds_to_standard_time(round(duration)),
                    })

                # Add the ad
                start_rounded = max(0, round(match["start_time"]))
                end_rounded = round(match["end_time"])
                duration_rounded = end_rounded - start_rounded
                all_rows.append({
                    "Brand": extract_brand_name(match["master_name"]),
                    "Description": os.path.splitext(match["master_name"])[0],
                    "Start Time (HH:MM:SS)": seconds_to_standard_time(start_rounded),
                    "End Time (HH:MM:SS)": seconds_to_standard_time(end_rounded),
                    "Ad Duration (HH:MM:SS)": seconds_to_standard_time(abs(duration_rounded)),
                })

                last_end = max(last_end, match['end_time'])

            # Check for a final gap after the last ad
            broadcast_duration = broadcast.duration if broadcast and broadcast.duration else last_end
            if broadcast_duration - last_end >= 1.0:
                gap_start = last_end
                gap_end = broadcast_duration
                duration = gap_end - gap_start
                all_rows.append({
                    "Brand": "Empty",
                    "Description": "Empty",
                    "Start Time (HH:MM:SS)": seconds_to_standard_time(round(gap_start)),
                    "End Time (HH:MM:SS)": seconds_to_standard_time(round(gap_end)),
                    "Ad Duration (HH:MM:SS)": seconds_to_standard_time(round(duration)),
                })

            if not all_rows:
                df = pd.DataFrame(columns=[
                    "Brand", "Description", "Start Time (HH:MM:SS)",
                    "End Time (HH:MM:SS)", "Ad Duration (HH:MM:SS)"
                ])
            else:
                df = pd.DataFrame(all_rows)

            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                from openpyxl.styles import PatternFill, Font, Alignment

                # Write dataframe to Excel, starting at row 4 (leaving row 3 blank)
                df.to_excel(writer, sheet_name="Ad Detection Results",
                            startrow=3, index=False)

                worksheet = writer.sheets["Ad Detection Results"]

                # Row 1: Radio Station
                worksheet.merge_cells('A1:E1')
                cell_station = worksheet['A1']
                cell_station.value = f"Radio Station: {broadcast.radio_station or 'N/A'}"
                cell_station.font = Font(bold=True, size=14)
                cell_station.alignment = Alignment(
                    horizontal="center", vertical="center")
                worksheet.row_dimensions[1].height = 20

                # Row 2: Broadcast Recording Filename
                worksheet.merge_cells('A2:E2')
                cell_filename = worksheet['A2']
                cell_filename.value = f"File: {broadcast.broadcast_recording or 'N/A'}"
                cell_filename.font = Font(bold=True, size=12)
                cell_filename.alignment = Alignment(
                    horizontal="center", vertical="center")
                worksheet.row_dimensions[2].height = 18

                # Style the dataframe header (now at row 4)
                yellow_fill = PatternFill(
                    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                )
                bold_font = Font(bold=True)
                center_alignment = Alignment(horizontal="center")

                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=4, column=col)
                    cell.fill = yellow_fill
                    cell.font = bold_font
                    cell.alignment = center_alignment

                worksheet.column_dimensions["A"].width = 20
                worksheet.column_dimensions["B"].width = 60
                worksheet.column_dimensions["C"].width = 18
                worksheet.column_dimensions["D"].width = 18
                worksheet.column_dimensions["E"].width = 20

            excel_data = output.getvalue()

            excel_filename = f"detection_results_{broadcast.broadcast_recording.replace('.mp3', '').replace('.wav', '')}.xlsx"

            session.query(ExcelReports).filter(
                ExcelReports.broadcast_id == broadcast_id
            ).delete()

            excel_record = ExcelReports(
                broadcast_id=broadcast_id,
                excel_data=excel_data,
                excel_filename=excel_filename,
                total_ads_detected=len(final_matches),
                file_size_bytes=len(excel_data),
            )
            session.add(excel_record)
            session.commit()

            print(
                f" Excel report stored in database for broadcast ID {broadcast_id}")

        except Exception as e:
            session.rollback()
            print(f" Error storing Excel: {e}")
            traceback.print_exc()
        finally:
            session.close()

    def get_radio_recordings_list(self):
        """Get list of all processed radio recordings"""
        session = self.Session()
        try:
            broadcasts = session.query(Broadcast.broadcast_recording).all()
            return [b[0] for b in broadcasts]
        except Exception as e:
            print(f"Error getting recordings list: {e}")
            return []
        finally:
            session.close()

    def download_excel_by_id(self, broadcast_id, save_path=None):
        """Download Excel file from database by radio filename"""
        session = self.Session()
        try:

            broadcast = (
                session.query(Broadcast).filter(
                    Broadcast.id == broadcast_id).first()
            )

            if not broadcast:
                print(f"No broadcast found for: {broadcast_id}")
                return None

            excel_record = (
                session.query(ExcelReports)
                .filter(ExcelReports.broadcast_id == broadcast_id)
                .first()
            )

            if not excel_record:
                print(f"No Excel report found for: {broadcast_id}")
                return None

            if save_path is None:
                save_path = excel_record.excel_filename

            with open(save_path, "wb") as f:
                f.write(excel_record.excel_data)

            print(f" Excel downloaded: {save_path}")
            print(f"  - Total ads detected: {excel_record.total_ads_detected}")
            print(f"  - File size: {excel_record.file_size_bytes} bytes")
            print(f"  - Created: {excel_record.created_timestamp}")

            return save_path

        except Exception as e:
            print(f" Error downloading Excel: {e}")
            return None
        finally:
            session.close()

    def stream_excel_report(self, broadcast_id):
        """Return Excel file as a FastAPI StreamingResponse. Generates report if missing and broadcast is processed."""
        session = self.Session()
        try:
            broadcast = (
                session.query(Broadcast).filter(
                    Broadcast.id == broadcast_id).first()
            )

            if not broadcast:
                raise HTTPException(
                    status_code=404, detail=f"Broadcast with ID {broadcast_id} not found.")

            excel_record = (
                session.query(ExcelReports)
                .filter(ExcelReports.broadcast_id == broadcast_id)
                .first()
            )

            if not excel_record:
                if broadcast.status == "Processed":
                    print(
                        f"Excel report for broadcast {broadcast_id} not found, but is processed. Generating report now.")

                    detection_results = (
                        session.query(AdDetectionResult)
                        .filter(AdDetectionResult.broadcast_id == broadcast_id)
                        .order_by(AdDetectionResult.start_time_seconds)
                        .all()
                    )

                    final_matches_reconstructed = []
                    for result in detection_results:
                        master_name = None
                        if result.clip_type == 'ad' and result.ad_id and result.ad_id != -1:
                            master_name = session.query(Ad.advertisement).filter(
                                Ad.id == result.ad_id).scalar()
                        else:
                            continue
                        # Ignoring songs for now
                        # elif result.clip_type == 'song' and result.ad_id and result.ad_id != -1:
                        #     master_name = session.query(Song.filename).filter(
                        #         Song.id == result.ad_id).scalar()

                        if not master_name:
                            master_name = f"{result.description or result.brand}.mp3"

                        final_matches_reconstructed.append({
                            "master_name": master_name,
                            "start_time": result.start_time_seconds,
                            "end_time": result.end_time_seconds,
                        })

                    self._generate_and_store_excel(
                        broadcast_id, final_matches_reconstructed)

                    excel_record = (
                        session.query(ExcelReports)
                        .filter(ExcelReports.broadcast_id == broadcast_id)
                        .first()
                    )

                    if not excel_record:
                        logging.error(
                            f"Failed to generate/retrieve Excel for broadcast {broadcast_id} after regeneration attempt.")
                        raise HTTPException(
                            status_code=500, detail="Failed to generate Excel report on-the-fly.")
                else:
                    raise HTTPException(
                        status_code=404, detail=f"Report for broadcast {broadcast_id} is not available (Status: {broadcast.status}).")

            file_stream = BytesIO(excel_record.excel_data)
            filename = (
                excel_record.excel_filename
                or f"Report_{broadcast.broadcast_recording}.xlsx"
            )

            return StreamingResponse(
                file_stream,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f"attachment; filename={filename}"},
            )

        except Exception as e:
            logging.error(
                f"Error streaming Excel for broadcast {broadcast_id}: {e}")
            logging.error(traceback.format_exc())
            if isinstance(e, HTTPException):
                raise e
            raise HTTPException(
                status_code=500, detail=f"An unexpected error occurred while generating the report.")
        finally:
            session.close()

    def get_detection_summary(self, radio_filename):
        """Get detailed summary of detection results"""
        session = self.Session()
        try:

            broadcast = (
                session.query(Broadcast)
                .filter(Broadcast.broadcast_recording == radio_filename)
                .first()
            )

            if not broadcast:
                print(f" No broadcast found for: {radio_filename}")
                return None

            results = (
                session.query(AdDetectionResult)
                .filter(AdDetectionResult.broadcast_id == broadcast.id)
                .order_by(AdDetectionResult.start_time_seconds.asc())
                .all()
            )

            if not results:
                print(f" No results found for: {radio_filename}")
                return None

            print(f"\n Detection Summary for: {radio_filename}")
            print("=" * 60)
            print(f"Total Ad Detected: {len(results)}")
            print(
                f"Processing Date: {results[0].detection_timestamp.strftime('%Y-%m-%d %H:%M:%S')}"
            )
            print(f"Broadcast ID: {results[0].broadcast_id}")
            print()

            brand_counts = {}
            total_duration = 0

            for result in results:
                brand_counts[result.brand] = brand_counts.get(
                    result.brand, 0) + 1
                total_duration += result.duration_seconds

            print(" Brand Breakdown:")
            for brand, count in sorted(
                brand_counts.items(), key=lambda x: x[1], reverse=True
            ):
                print(f"  {brand}: {count} ads")

            print(
                f"\n Total Ad Duration: {seconds_to_standard_time(total_duration)}")
            print(
                f" Average Correlation Score: {np.mean([r.correlation_score for r in results]):.4f}"
            )

            print(f"\n First 5 Detections (with IDs):")
            for i, result in enumerate(results[:5], 1):
                print(
                    f"  {i}. {result.brand} | Ad ID: {result.ad_id} | {seconds_to_standard_time(result.start_time_seconds)}"
                )

            return {
                "total_ads": len(results),
                "brands": brand_counts,
                "total_duration": total_duration,
                "avg_correlation": np.mean([r.correlation_score for r in results]),
                "broadcast_id": results[0].broadcast_id,
            }

        except Exception as e:
            print(f"Error getting summary: {e}")
            return None
        finally:
            session.close()

    def extract_clip_from_broadcast(
        self,
        broadcast_id,
        brand_or_artist,
        advertisement_or_name,
        clip_type,
        start_time,
        end_time,
        ad_masters_folder="ad_masters",
        songs_folder="songs",
    ):
        session = self.Session()
        start_time = time.time()
        try:

            broadcast = (
                session.query(Broadcast).filter(
                    Broadcast.id == broadcast_id).first()
            )

            if not broadcast:
                print(f" Broadcast with ID {broadcast_id} not found")
                return False

            broadcast_file = broadcast.broadcast_recording
            print(f" Processing broadcast: {broadcast_file}")

            if clip_type.lower() == "speech":
                print(" Speech type detected - inserting directly to database")

                speech_result = AdDetectionResult(
                    brand="Speech",
                    description="",
                    start_time_seconds=float(start_time),
                    end_time_seconds=float(end_time),
                    duration_seconds=float(end_time - start_time),
                    correlation_score=1.0,
                    raw_correlation=1.0,
                    mfcc_correlation=1.0,
                    overlap_duration=float(end_time - start_time),
                    total_matches_found=1,
                    ad_id=-1,
                    broadcast_id=broadcast_id,
                    processing_status="manual_entry",
                    clip_type="speech",
                )

                session.add(speech_result)
                session.commit()
                print(f" Speech entry added to detection results")
                return True

            radio_file_path = get_first_file_path("radio_recording")
            broadcast_audio, broadcast_sr = load_audio(radio_file_path)
            if broadcast_audio is None:
                print(f" Failed to load broadcast audio")
                return False
            else:
                print("Loaded audio file")

            start_sample = int(start_time * broadcast_sr)
            end_sample = int(end_time * broadcast_sr)

            if (
                start_sample < 0
                or end_sample > len(broadcast_audio)
                or start_sample >= end_sample
            ):
                print(f" Invalid time range: {start_time}s to {end_time}s")
                print(
                    f"   Broadcast duration: {len(broadcast_audio)/broadcast_sr:.2f}s"
                )
                return False

            extracted_clip = broadcast_audio[start_sample:end_sample]
            duration = len(extracted_clip) / broadcast_sr

            print(
                f" Extracted clip: {duration:.2f}s ({start_time}s - {end_time}s)")

            if clip_type.lower() == "ad":

                safe_brand = "".join(
                    c for c in brand_or_artist if c.isalnum() or c in (" ", "-", "_")
                ).rstrip()
                safe_desc = "".join(
                    c
                    for c in advertisement_or_name
                    if c.isalnum() or c in (" ", "-", "_")
                ).rstrip()
                filename = f"{safe_brand}_{safe_desc}.mp3"
                output_folder = ad_masters_folder

            elif clip_type.lower() == "song":

                safe_artist = "".join(
                    c for c in brand_or_artist if c.isalnum() or c in (" ", "-", "_")
                ).rstrip()
                safe_song = "".join(
                    c
                    for c in advertisement_or_name
                    if c.isalnum() or c in (" ", "-", "_")
                ).rstrip()
                filename = f"{safe_artist}_{safe_song}.mp3"
                output_folder = songs_folder

            else:
                print(
                    f" Invalid clip type: {clip_type}. Use 'ad', 'song', or 'speech'")
                return False

            os.makedirs(output_folder, exist_ok=True)
            output_path = os.path.join(output_folder, filename)

            try:

                temp_wav = io.BytesIO()
                sf.write(temp_wav, extracted_clip, broadcast_sr, format="WAV")
                temp_wav.seek(0)

                audio_segment = AudioSegment.from_wav(temp_wav)
                audio_segment.export(output_path, format="mp3", bitrate="192k")
                print(f" Saved {clip_type} clip: {output_path}")
            except Exception as e:
                print(f" Error saving audio file: {e}")
                return False

            if clip_type.lower() == "ad":

                new_ad = Ad(
                    brand=brand_or_artist,
                    advertisement=filename,
                    duration=int(duration),
                    status="Active",
                )
                session.add(new_ad)
                session.commit()

                ad_id = new_ad.id
                print(f" Added to ads table (ID: {ad_id})")

                detection_result = AdDetectionResult(
                    brand=brand_or_artist,
                    description=advertisement_or_name,
                    start_time_seconds=float(start_time),
                    end_time_seconds=float(end_time),
                    duration_seconds=duration,
                    correlation_score=1.0,
                    raw_correlation=1.0,
                    mfcc_correlation=1.0,
                    overlap_duration=duration,
                    total_matches_found=1,
                    ad_id=ad_id,
                    broadcast_id=broadcast_id,
                    processing_status="manual_extraction",
                    clip_type="ad",
                )
                session.add(detection_result)

            elif clip_type.lower() == "song":

                new_song = Song(
                    artist=brand_or_artist,
                    name=advertisement_or_name,
                    filename=filename,
                    duration=int(duration),
                    upload_date=datetime.now(),
                    status="Active",
                )
                session.add(new_song)
                session.commit()
                song_id = new_song.id
                print(f" Added to songs table: {filename}")
                song_detection_result = AdDetectionResult(
                    brand=brand_or_artist,
                    description=advertisement_or_name,
                    start_time_seconds=float(start_time),
                    end_time_seconds=float(end_time),
                    duration_seconds=duration,
                    correlation_score=1.0,
                    raw_correlation=1.0,
                    mfcc_correlation=1.0,
                    overlap_duration=duration,
                    total_matches_found=1,
                    ad_id=song_id,
                    broadcast_id=broadcast_id,
                    processing_status="manual_extraction",
                    clip_type="song",
                )
                session.add(song_detection_result)
            session.commit()

            elapsed = time.time() - start_time
            logging.info(
                f"extract_clip_from_broadcast completed in {elapsed:.2f} seconds"
            )

            print(f" Successfully processed {clip_type}: {filename}")
            print(f"    Saved to: {output_folder}")
            print(f"    Duration: {duration:.2f}s")
            print(f"    Linked to broadcast ID: {broadcast_id}")

            return True

        except Exception as e:
            session.rollback()
            print(f" Error extracting clip: {e}")
            error_msg = f"extract_clip_from_broadcast failed after {time.time() - start_time:.2f} seconds: {str(e)}"
            logging.error(error_msg)
            logging.error(traceback.format_exc())
            raise HTTPException(status_code=505, detail=str(e))

            return False
        finally:
            session.close()


def get_first_file_path(folder_path: str) -> str | None:
    try:
        files = os.listdir(folder_path)
        files = [f for f in files if os.path.isfile(
            os.path.join(folder_path, f))]
        if not files:
            return None
        first_file = files[0]
        return os.path.abspath(os.path.join(folder_path, first_file))
    except FileNotFoundError:
        return None


def process_single_radio_clip(
    broadcast_id,
    ad_masters_folder="ad_masters",
    radio_recording_directory="radio_recording",
    correlation_threshold=0.65,
    db_manager=None,
):
    """Process a single radio recording clip and save everything to database"""

    start_time = time.time()

    try:
        if db_manager is None:
            db_manager = EnhancedRadioRecordingManager()

        session = db_manager.Session()
        try:
            broadcast = (
                session.query(Broadcast).filter(
                    Broadcast.id == broadcast_id).first()
            )
            if broadcast:
                broadcast.status = "Processing"
                session.commit()
                print(
                    f" Set status to Processing for broadcast {broadcast_id}")
        except Exception as e:
            print(f"Error setting status: {e}")
        finally:
            session.close()

        radio_file_path = get_first_file_path(radio_recording_directory)
        print(f"\n Processing: {radio_file_path}")
        print("=" * 50)

        print(" Loading advertisement masters...")
        masters = {}
        for filename in os.listdir(ad_masters_folder):
            if filename.endswith((".wav", ".mp3")):
                filepath = os.path.join(ad_masters_folder, filename)
                audio, sr = load_audio(filepath)
                if audio is not None:
                    masters[filename] = {
                        "audio": audio,
                        "sr": sr,
                        "duration": len(audio) / sr,
                    }

        print(f" Loaded {len(masters)} advertisement masters")

        print(" Loading radio recording...")
        radio_recording, radio_sr = load_audio(radio_file_path)
        if radio_recording is None:
            print(f" Error: Could not load radio recording")
            db_manager.update_broadcast_status(broadcast_id, "Failed")
            return False

        radio_duration = len(radio_recording) / radio_sr
        print(
            f" Loaded radio recording (Duration: {seconds_to_standard_time(radio_duration)})"
        )

        print(" Finding advertisement matches...")
        all_matches = {}
        total_matches = 0

        for master_name, master_data in masters.items():
            matches = find_matches_improved(
                master_data["audio"],
                master_data["sr"],
                radio_recording,
                radio_sr,
                threshold=correlation_threshold,
            )
            all_matches[master_name] = matches
            total_matches += len(matches)
            if len(matches) > 0:
                print(f"{master_name}: {len(matches)} matches")

        print(f"Total raw matches found: {total_matches}")

        final_matches = db_manager.save_detection_results(
            broadcast_id, all_matches, radio_file_path
        )

        session = db_manager.Session()
        try:
            broadcast = (
                session.query(Broadcast).filter(
                    Broadcast.id == broadcast_id).first()
            )
            if broadcast:
                broadcast.processing_time = datetime.now()
                broadcast.status = "Processed"
                session.commit()
                print(
                    f" Set processing_time for broadcast {broadcast_id} AFTER completion"
                )
        except Exception as e:
            print(f" Error setting final processing_time: {e}")
        finally:
            session.close()

        elapsed = time.time() - start_time
        logging.info(
            f"process_single_radio_clip completed in {elapsed:.2f} seconds")

        if final_matches > 0:
            print(f" Processing successful! Final matches: {final_matches}")
            return True
        else:
            print("o matches found above threshold")

            session = db_manager.Session()
            try:
                broadcast = (
                    session.query(Broadcast)
                    .filter(Broadcast.id == broadcast_id)
                    .first()
                )
                if broadcast and broadcast.status != "Processed":
                    broadcast.status = "No Matches"
                    session.commit()
            finally:
                session.close()
            return False

    except Exception as e:
        elapsed = time.time() - start_time
        error_msg = (
            f"process_single_radio_clip failed after {elapsed:.2f} seconds: {str(e)}"
        )
        print(error_msg)
        print(traceback.format_exc())

        if db_manager:
            db_manager.update_broadcast_status(broadcast_id, "Failed")

        raise HTTPException(status_code=500, detail=str(e))


def fetch_excel_report(broadcast_id, download_path=None):
    """Fetch Excel report from database by radio filename"""
    db_manager = EnhancedRadioRecordingManager()
    return db_manager.stream_excel_report(broadcast_id)


def get_report_summary(radio_filename):
    """Get detailed summary of a specific report"""
    db_manager = EnhancedRadioRecordingManager()
    return db_manager.get_detection_summary(radio_filename)


def download_report_by_radio_name(radio_filename, save_folder="downloads"):
    """Download Excel report by radio filename"""
    os.makedirs(save_folder, exist_ok=True)
    save_path = os.path.join(
        save_folder,
        f"report_{radio_filename.replace('.mp3', '').replace('.wav', '')}.xlsx",
    )
    return fetch_excel_report(radio_filename, save_path)


def extract_clip(
    broadcast_id,
    brand_or_artist,
    advertisement_or_name,
    clip_type,
    start_time,
    end_time,
):
    db_manager = EnhancedRadioRecordingManager()
    return db_manager.extract_clip_from_broadcast(
        broadcast_id,
        brand_or_artist,
        advertisement_or_name,
        clip_type,
        start_time,
        end_time,
    )


def reprocess_broadcast(broadcast_id):
    db_manager = EnhancedRadioRecordingManager()
    session = db_manager.Session()
    try:
        broadcast = get_broadcast(session, broadcast_id)
        # Check for ads added after last broadcast processing
        newer_ads = (
            session.query(Ad)
            .filter(
                Ad.upload_date > broadcast.processing_time,
                Ad.status == "Active"
            )
            .all()
        )

        if not newer_ads:
            err_msg = f" No new ads since {broadcast.processing_time.strftime('%Y-%m-%d %H:%M:%S')}"
            print(err_msg)
            return {"message": err_msg}

        masters = {}
        for ad in newer_ads:
            filepath = os.path.join("ad_masters", ad.advertisement)
            if os.path.exists(filepath):
                audio, sr = load_audio(filepath)
                if audio is not None:
                    masters[ad.advertisement] = {
                        "audio": audio,
                        "sr": sr,
                        "duration": len(audio) / sr,
                    }

        if not masters:
            print("No valid new ad masters found")
            return False

        # Download broadcast
        local_path = download_file_from_s3(
            broadcast.filename,
            "broadcasts",
            s3_folder="broadcasts"
        )
        radio_audio, radio_sr = load_audio(local_path)
        if radio_audio is None:
            print(" Could not load broadcast audio")
            return False

        all_matches = {}
        total_matches = 0

        # Look for ads in the recording
        for master_name, master_data in masters.items():
            matches = find_matches_improved(
                master_data["audio"],
                master_data["sr"],
                radio_audio,
                radio_sr,
                threshold=0.65,
            )
            all_matches[master_name] = matches
            total_matches += len(matches)

            if len(matches) > 0:
                print(f" {master_name}: {len(matches)} matches")

        # Save detection results
        if total_matches > 0:
            new_matches = db_manager.save_detection_results(
                broadcast.id, all_matches, local_path
            )

            session = db_manager.Session()
            try:
                broadcast_obj = (
                    session.query(Broadcast)
                    .filter(Broadcast.id == broadcast.id)
                    .first()
                )
                if broadcast_obj:
                    broadcast_obj.processing_time = datetime.now()
                    broadcast_obj.status = "Processed"
                    session.commit()
                    print(
                        f" Updated processing_time for broadcast {broadcast.id}"
                    )
            except Exception as e:
                print(f" Error updating processing_time: {e}")
            finally:
                session.close()

            print(
                f" Added {new_matches} new detections to broadcast {broadcast.id}"
            )
        else:
            print(" No new matches found")

        return True

    except Exception as e:
        print(f" Error processing {broadcast.filename}: {e}")
        print(f"Full traceback: {traceback.format_exc()}")
        return False

    finally:
        db_manager.update_broadcast_status(broadcast_id, "Processed")


def batch_reprocess_broadcasts(db_manager=None, max_workers=4):
    if db_manager is None:
        db_manager = EnhancedRadioRecordingManager()

    session = db_manager.Session()
    try:
        processed_broadcasts = (
            session.query(Broadcast).filter(
                Broadcast.status == "Processed"
            ).all()
        )

        print(
            f"Found {len(processed_broadcasts)} processed broadcasts to check")

        broadcasts_to_process = []

        for broadcast in processed_broadcasts:
            newer_ads = (
                session.query(Ad)
                .filter(
                    Ad.upload_date > broadcast.processing_time, Ad.status == "Active"
                )
                .all()
            )

            if not newer_ads:
                print(
                    f" No new ads since {broadcast.processing_time.strftime('%Y-%m-%d %H:%M:%S')}"
                )
                continue

            broadcasts_to_process.append(
                {"broadcast": broadcast, "new_ads": newer_ads})

        if not broadcasts_to_process:
            print("No broadcasts require reprocessing")
            return

        print(
            f" Starting reprocessing of {len(broadcasts_to_process)} broadcasts..."
        )

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {}

            for item in broadcasts_to_process:
                future = executor.submit(
                    reprocess_broadcast,
                    db_manager,
                    item["broadcast"],
                    item["new_ads"],
                )
                futures[future] = item["broadcast"].id

            for future in as_completed(futures):
                broadcast_id = futures[future]
                try:
                    success = future.result()
                    if success:
                        print(
                            f" Broadcast {broadcast_id}: Successfully processed"
                        )
                    else:
                        print(f" Broadcast {broadcast_id}: Processing failed")
                except Exception as e:
                    print(f" Broadcast {broadcast_id}: Error - {e}")

        print("Periodic check completed!")

    except Exception as e:
        print(f" Error during periodic check: {e}")
    finally:
        session.close()


def setup_weekly_scheduler():
    """Setup the weekly scheduler for checking new ads"""

    def run_weekly_check():
        print(" Running scheduled weekly check for new ads...")
        try:
            batch_reprocess_broadcasts()
        except Exception as e:
            print(f" Weekly check failed: {e}")

    schedule.every().monday.at("03:00").do(run_weekly_check)

    print(" Weekly scheduler setup complete. Will run every Monday at 3:00 AM")

    def run_scheduler():
        while True:
            schedule.run_pending()
            time.sleep(60)

    scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
    scheduler_thread.start()

    return scheduler_thread


def main():
    """Main function to test the enhanced system"""
    print(" Starting Enhanced Radio Ad Detection System")
    print("=" * 60)

    db_manager = EnhancedRadioRecordingManager()

    scheduler_thread = setup_weekly_scheduler()

    print(" Running immediate check for new ads...")

    print(" System startup completed!")
    print(" Weekly scheduler is running in background")

    return db_manager, scheduler_thread


if __name__ == "__main__":
    pass
    # main()
